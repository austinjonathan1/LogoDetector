# Current Class #: 15
# VCCI_Japan
# CE_Europe
# WEEE_Europe
# CSA_UL_US_Canada
# RCM_ Australia
# RoHs_China
# KC_Korea
# RoHs_Taiwan
# CCC_China
# EAC_Russia
# Cp_Morocco
# Anatel_Brazil
# UKCA_UK
# BIS_India
# NOM_Mexico
# LIBRARIES, DECLARATIONS, & FLAGS
total_number_of_logos = 15
import time
from absl import app, flags, logging
from absl.flags import FLAGS
import cv2
import numpy as np
import tensorflow as tf
from yolov3_tf2.models import (YoloV3, YoloV3Tiny)
from yolov3_tf2.dataset import transform_images, load_tfrecord_dataset
from yolov3_tf2.utils import draw_outputs
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pdf2image import convert_from_path, convert_from_bytes
from pdf2image.exceptions import (PDFInfoNotInstalledError, PDFPageCountError, PDFSyntaxError)
extola = {} # excel to label conversion dictionary
extola["ALL CE COUNTRIES"] = ["CE_Europe", "WEEE_Europe"]
extola["AUSTRALIA"] = ["RCM_ Australia"]
extola["BRAZIL"] = ["Anatel_Brazil"]
extola["CANADA"] = ["CSA_UL_US_Canada"]
extola["CHINA"] = ["CCC_China", "RoHs_China"]
extola["INDIA"] = ["BIS_India"]
extola["JAPAN"] = ["VCCI_Japan"]
extola["KOREA, REPUBLIC OF"] = ["KC_Korea"]
extola["KOREA, DEMOCRATIC PEOPLE'S REPUBLIC OF"] = ["KC_Korea"]
extola["MEXICO"] = ["NOM_Mexico"]
extola["MOROCCO"] = ["Cp_Morocco"]
extola["RUSSIAN FEDERATION"] = ["EAC_Russia"]
extola["TAIWAN"] = ["RoHs_Taiwan"]
extola["UNITED STATES"] = ["CSA_UL_US_Canada"]
extola["UNITED KINGDOM"] = ["UKCA_UK"]
flags.DEFINE_integer('count', 1, 'number of label/excel pairs')
# WEIGHTS, CLASSES, & FLAGS ROUTING
def main(_argv):
    physical_devices = tf.config.experimental.list_physical_devices('GPU')
    if len(physical_devices) > 0:
        tf.config.experimental.set_memory_growth(physical_devices[0], True)
    yolo = YoloV3(classes=total_number_of_logos) # number of classes/logos, needs to be updated if another logo is added
    yolo.load_weights('./weights/yolov3-custom.tf').expect_partial() # file path to weights
    class_names = [c.strip() for c in open('./data/labels/custom.names').readlines()] # file path to classes list, needs to be updated if another logo is added
    if FLAGS.count:
        count = FLAGS.count
    excel = []
    images = []
    for i in range(count):
        con = convert_from_path('data/pdf/test (' + str(i+1) + ').pdf', output_folder='data/images', fmt="jpg", single_file=True, output_file='test (' + str(i+1) + ')')
        excel.append('data/excel/test (' + str(i+1) + ').xlsx')
        images.append('data/images/test (' + str(i+1) + ').jpg')
    raw_images = []
    for image in images:
        img_raw = tf.image.decode_image(
            open(image, 'rb').read(), channels=3)
        raw_images.append(img_raw)
    i = 0 # index number for main loop
    logos = [] # list of detected logos for each image
    approvals = [] # list of excel data for each image
    for raw_img in raw_images:
        img = tf.expand_dims(raw_img, 0)
        img = transform_images(img, 416) # image size
        t1 = time.time()
        boxes, scores, classes, nums = yolo(img)
        t2 = time.time()
        logging.info('time: {}'.format(t2 - t1))
        img = cv2.cvtColor(raw_img.numpy(), cv2.COLOR_RGB2BGR)
        img = draw_outputs(img, (boxes, scores, classes, nums), class_names)
        cv2.imwrite('./detections/detection (' + str(i+1) + ').jpg', img) # image output
# LABEL EXTRACTION
        temp_names = [] # temporary list for each image's logo detections
        for j in range(nums[0]):
            repeat = True
            temp_pair = [] # temporary list for each logo and its status
            if (j > 0):
                for k in range(len(temp_names)):
                    if (class_names[int(classes[0][j])] == temp_names[k][0]):
                        repeat = False
                        break
            if (repeat): # if not a repeated logo, update main logo list
                temp_pair.append(class_names[int(classes[0][j])]) # append logo 
                temp_pair.append(False) # append status
                temp_names.append(temp_pair) # append pair
        logos.append(temp_names) # append names list to main logo list
# EXCEL EXTRACTION
        wb = load_workbook(excel[i])
        sheet = wb.active
        rows = sheet.max_row
        temp_sheet = [] # temporary list for each image's excel data
        for j in range(rows-1):
            temp_rows = [] # temporary list for each row's excel data
            temp_rows.append(str(sheet.cell(row=j+2, column=4).value).upper().strip())
            temp_rows.append(str(sheet.cell(row=j+2, column=5).value).upper().strip())
            temp_rows.append("00FF0000") # Red by default
            temp_sheet.append(temp_rows)
        approvals.append(temp_sheet) # append sheet list to main approvals list
# EXCEL TRANSLATION
        for j in range(len(approvals[i])):
            if (approvals[i][j][0] in extola):
                temp_trans = extola[approvals[i][j][0]]
            else:
                temp_trans = ["NAL"] # No Associated Logo
            approvals[i][j][0] = temp_trans
# EXCEL COMPARED TO LABEL
# "APPROVAL STATUS"             "On label"   "Not on label"
# "APPROVED"                    "Green"     "Red"
# "NO REQUIREMENTS"             "Red"       "Green"
# "APPROVAL NOT APPLICABLE"     "Red"       "Green"
# "APPROVAL NOT REQUIRED"       "Red"       "Green"
# "CONTACT CISCO PARTNER/IOR"   "Red"       "Green"
# "NOT APPROVED"                "Red"       "Green"
# "PENDING"                     "Red"       "Green"
# "RENEWAL IN PROGESS"          "Red"       "Green"
# "NONE"/"UNKNOWN"              "Red"       "Red"
# 
# "00FF0000" (Red) needs attention
# "0000FF00" (Green) good to go
# "000000FF" (Blue) missing logo
#         
        for j in range(len(approvals[i])):
            flag = True
            k = 0
            temp_count = 0
            while (flag):
                if (k == len(logos[i])): # logo not on label
                    flag = False
                    if (approvals[i][j][1] == "APPROVED"):
                        approvals[i][j][2] = "00FF0000" # Red
                    elif ((approvals[i][j][1] == "APPROVAL NOT APPLICABLE")or(approvals[i][j][1] == "APPROVAL NOT REQUIRED")or(approvals[i][j][1] == "CONTACT CISCO PARTNER/IOR")or
                          (approvals[i][j][1] == "NOT APPROVED")or(approvals[i][j][1] == "PENDING")or(approvals[i][j][1] == "RENEWAL IN PROGESS")or(approvals[i][j][1] == "NO REQUIREMENTS")):
                        approvals[i][j][2] = "0000FF00" # Green
                    elif ((approvals[i][j][1] == "NONE")or(approvals[i][j][1] == "UNKNOWN")):
                        approvals[i][j][2] = "00FF0000" # Red
                        sheet.cell(row=j+2, column=5).value = "Unknown"
                elif (approvals[i][j][0][0] == "NAL"): # no logo to detect
                    flag = False
                    if ((approvals[i][j][1] == "APPROVAL NOT APPLICABLE")or(approvals[i][j][1] == "APPROVAL NOT REQUIRED")or(approvals[i][j][1] == "CONTACT CISCO PARTNER/IOR")or
                        (approvals[i][j][1] == "NOT APPROVED")or(approvals[i][j][1] == "PENDING")or(approvals[i][j][1] == "RENEWAL IN PROGESS")or(approvals[i][j][1] == "APPROVED")or(approvals[i][j][1] == "NO REQUIREMENTS")):
                        approvals[i][j][2] = "0000FF00" # Green
                    elif ((approvals[i][j][1] == "NONE")or(approvals[i][j][1] == "UNKNOWN")):
                        approvals[i][j][2] = "00FF0000" # Red
                        sheet.cell(row=j+2, column=5).value = "Unknown"  
                else: # continue or logo on label
                    for X in range(len(approvals[i][j][0])):
                        if (approvals[i][j][0][X] == logos[i][k][0]): # logo on label
                            logos[i][k][1] = True
                            temp_count+=1
                            if (temp_count == len(approvals[i][j][0])):
                                flag = False
                            if ((approvals[i][j][1] == "APPROVAL NOT APPLICABLE")or(approvals[i][j][1] == "APPROVAL NOT REQUIRED")or(approvals[i][j][1] == "CONTACT CISCO PARTNER/IOR")or
                                (approvals[i][j][1] == "NOT APPROVED")or(approvals[i][j][1] == "PENDING")or(approvals[i][j][1] == "RENEWAL IN PROGESS")):
                                approvals[i][j][2] = "00FF0000" # Red
                            elif ((temp_count == len(approvals[i][j][0]))and(approvals[i][j][1] == "APPROVED")or(approvals[i][j][1] == "NO REQUIREMENTS")):
                                approvals[i][j][2] = "0000FF00" # Green
                            elif ((approvals[i][j][1] == "NONE")or(approvals[i][j][1] == "UNKNOWN")):
                                approvals[i][j][2] = "00FF0000" # Red
                                sheet.cell(row=j+2, column=5).value = "Unknown"
                k+=1
            sheet.cell(row=j+2, column=5).fill = PatternFill(start_color=approvals[i][j][2], end_color=approvals[i][j][2], fill_type='solid')
# LABEL COMPARED TO EXCEL
        new_row=1
        for j in range(len(logos[i])):
            if (logos[i][j][1] == False): # not on excel so add it in a new row
                sheet.cell(row=new_row+rows, column=1).value = str(sheet.cell(row=rows, column=1).value) #1 Product Name
                sheet.cell(row=new_row+rows, column=3).value = str(sheet.cell(row=rows, column=3).value) #3 Desc
                sheet.cell(row=new_row+rows, column=4).value = logos[i][j][0] #4 Country
                sheet.cell(row=new_row+rows, column=5).value = "Unknown" #5 Approval Status
                sheet.cell(row=new_row+rows, column=5).fill = PatternFill(start_color="000000FF", end_color="000000FF", fill_type='solid') #5 Blue
                for k in range(5):
                    sheet.cell(row=new_row+rows, column=k+6).value = str(sheet.cell(row=rows, column=k+6).value) #6-10
                new_row+=1
        wb.save(excel[i])
        i+=1
# DISPLAY
    for j in range(i):
        print("\nL" + str(j+1) + ": ", end="")
        temp_print = []
        for k in range(len(logos[j])):
            temp_print.append(logos[j][k][0])
        print(temp_print, "\nE" + str(j+1) + ": ", end="")
        temp_print = []
        for k in range(len(approvals[j])):
            temp_print.append(approvals[j][k][0])
        print(temp_print)
    print("")
# MAIN
if __name__ == '__main__':
    try:
        app.run(main)
    except SystemExit:
        pass
